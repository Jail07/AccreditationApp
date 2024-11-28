import sys
import re
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem
from PyQt5.QtGui import QColor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class AccreditationApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Обработка аккредитации')

        layout = QVBoxLayout()

        self.tableWidget = QTableWidget(self)
        layout.addWidget(self.tableWidget)

        self.loadButton = QPushButton('Загрузить таблицу', self)
        self.loadButton.clicked.connect(self.loadFile)
        layout.addWidget(self.loadButton)

        self.processButton = QPushButton('Обработать данные', self)
        self.processButton.clicked.connect(self.processData)
        layout.addWidget(self.processButton)

        self.saveButton = QPushButton('Сохранить таблицу', self)
        self.saveButton.clicked.connect(self.saveFile)
        layout.addWidget(self.saveButton)

        self.setLayout(layout)
        self.df = None  # Для хранения данных
        self.show()

    def loadFile(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Загрузить файл", "", "Excel Files (*.xls *.xlsx)",
                                                  options=options)

        if fileName:
            self.df = pd.read_excel(fileName)
            self.displayTable()


    # 1989-09-14 00:00:00
    # № п/п
    # 1.0 2.0


    def displayTable(self):
        self.tableWidget.setRowCount(len(self.df))
        self.tableWidget.setColumnCount(len(self.df.columns))
        self.tableWidget.setHorizontalHeaderLabels(self.df.columns)

        for i in range(len(self.df)):
            for j in range(len(self.df.columns)):
                item = QTableWidgetItem(str(self.df.iloc[i, j]))
                self.tableWidget.setItem(i, j, item)

    def processData(self):
        if self.df is None:
            return

        # Фильтруем и очищаем данные
        self.df = self.cleanData(self.df)

        # Проверяем, если ключевые столбцы существуют
        required_columns = ['Фамилия', 'Имя', 'Дата рождения']
        if all(col in self.df.columns for col in required_columns):
            # Исключаем строки с пустыми значениями в ключевых столбцах
            valid_rows = self.df.dropna(subset=required_columns)

            # Проверка повторов только для валидных строк
            self.df['Повтор'] = self.df.index.isin(
                valid_rows[valid_rows.duplicated(subset=required_columns, keep=False)].index
            )

            # Обновляем отображение с выделением
            for i in range(len(self.df)):
                row_data = self.df.iloc[i]
                is_row_empty_except_id = (
                    all(pd.isna(row_data[col]) for col in self.df.columns if col != '№ п/п') and
                    not pd.isna(row_data.get('№ п/п', None))
                )
                for j in range(len(self.df.columns)):
                    item = QTableWidgetItem(str(self.df.iloc[i, j]))
                    if not is_row_empty_except_id:
                        if self.df.iloc[i].get('Повтор', False):  # Если повтор, выделяем цветом
                            item.setBackground(QColor(255, 0, 0))  # Красный цвет
                        elif pd.isna(self.df.iloc[i, j]) and j != 0:  # Если отсутствуют данные
                            item.setBackground(QColor(255, 165, 0))  # Оранжевый цвет
                    self.tableWidget.setItem(i, j, item)
        else:
            print("Не хватает столбцов для проверки повторов")

    def removeExtraSpaces(self, value):
        """Удаляет лишние пробелы, неразрывные пробелы и нормализует строку"""
        value = value.replace('\u00A0', ' ')  # Заменяет неразрывные пробелы на обычные
        value = value.strip()  # Убирает пробелы с начала и конца
        value = re.sub(r'\s+', ' ', value)  # Заменяет несколько пробелов на один
        return value

    def cleanData(self, df):
        # Функция для очистки данных
        for col in df.columns:
            if col == '№ п/п':
                # Убираем преобразование чисел в float
                df[col] = df[col].apply(lambda x: int(x) if pd.notna(x) and str(x).isdigit() else x)
            elif col == 'Дата рождения':
                # Очистка и нормализация даты рождения
                df[col] = df[col].apply(lambda x: self.normalizeDate(x) if pd.notna(x) else None)
            else:
                # Очистка пробелов для остальных столбцов
                df[col] = df[col].apply(lambda x: self.removeExtraSpaces(str(x)) if pd.notna(x) else None)

        return df

    def normalizeDate(self, date_str):
        """Нормализует дату рождения, убирает ненужные символы и форматирует"""
        if pd.isna(date_str) or date_str == '':
            return None  # Оставляем пустую ячейку, если изначально пустая

        # Удаляем все лишние символы, кроме разделителей дат
        date_str = re.sub(r'[^\d.]', '', str(date_str)).strip()

        try:
            # Пробуем распознать дату
            date_obj = pd.to_datetime(date_str, errors='coerce', dayfirst=True)
            if pd.notna(date_obj):
                # Преобразуем в формат "дд.мм.гггг"
                formatted_date = date_obj.strftime('%d.%m.%Y')

                birth_year = date_obj.year
                if birth_year < 1950:
                    return None
                return formatted_date
            else:
                return None  # Если дата невалидна, возвращаем None
        except Exception as e:
            print(f"Ошибка обработки даты: {date_str} - {e}")
            return None  # Если что-то пошло не так, возвращаем None

    def saveFile(self):
        if self.df is None:
            return

        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)",
                                                  options=options)

        if fileName:
            # Создаём копию данных без столбца "Повтор"
            df_to_save = self.df.drop(columns=['Повтор'], errors='ignore')

            # Создаём Excel-файл
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Обработанные данные"

            # Записываем данные
            for row in dataframe_to_rows(df_to_save, index=False, header=True):
                sheet.append(row)

            # Окрашиваем ячейки в Excel
            for i, row in enumerate(self.df.itertuples(index=False), start=2):
                for j, col in enumerate(self.df.columns, start=1):
                    cell = sheet.cell(row=i, column=j)
                    if getattr(row, 'Повтор', False):  # Если повтор
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            wb.save(fileName)
            print(f"Файл сохранён: {fileName}")


app = QApplication(sys.argv)
ex = AccreditationApp()
sys.exit(app.exec_())
