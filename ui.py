from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, QTextEdit, \
    QHBoxLayout
from PyQt5.QtGui import QColor
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from data_processing import DataProcessor
from file_manager import FileManager
import pandas as pd


class AccreditationApp(QWidget):
    def __init__(self, db_manager):
        super().__init__()
        self.db_manager = db_manager
        self.initUI()

    def logMessage(self, message):
        """
        Добавляет сообщение в текстовое поле логов.
        Если текстовое поле скрыто, оно становится видимым.
        """
        if not self.logText.isVisible():
            self.logText.show()
        self.logText.append(message)

    def initUI(self):
        self.setWindowTitle('Обработка аккредитации')

        # Основной макет
        layout = QHBoxLayout()

        # Левая часть с таблицей и кнопками
        left_layout = QVBoxLayout()

        # Таблица
        self.tableWidget = QTableWidget(self)
        left_layout.addWidget(self.tableWidget)

        # Кнопки
        self.loadButton = QPushButton('Загрузить таблицу', self)
        self.loadButton.clicked.connect(self.loadFile)
        left_layout.addWidget(self.loadButton)

        self.checkButton = QPushButton('Проверка данных', self)
        self.checkButton.clicked.connect(self.checkData)
        left_layout.addWidget(self.checkButton)

        self.uploadButton = QPushButton('Загрузить в БД', self)
        self.uploadButton.setEnabled(False)  # Кнопка неактивна, пока не выполнена проверка данных
        self.uploadButton.clicked.connect(self.uploadToDB)
        left_layout.addWidget(self.uploadButton)

        self.saveButton = QPushButton('Сохранить файл', self)
        self.saveButton.clicked.connect(self.saveFile)
        left_layout.addWidget(self.saveButton)

        self.blacklistButton = QPushButton('Добавить в черный список', self)
        self.blacklistButton.clicked.connect(self.addToBlacklist)
        left_layout.addWidget(self.blacklistButton)

        layout.addLayout(left_layout)

        # Правая часть с логами
        self.logText = QTextEdit(self)
        self.logText.setReadOnly(True)  # Поле только для чтения
        self.logText.hide()  # Скрываем поле до нажатия на "Проверка данных"
        layout.addWidget(self.logText)

        self.setLayout(layout)
        self.df = None  # Для хранения данных
        self.processor = DataProcessor()
        self.file_manager = FileManager()

    def loadFile(self):
        file_name = self.file_manager.openFile(self)
        if file_name:
            self.df = pd.read_excel(file_name)
            self.displayTable()
            self.logText.clear()  # Очищаем логи при загрузке нового файла

    def displayTable(self):
        """
        Отображает содержимое DataFrame в QTableWidget.
        """
        self.tableWidget.setRowCount(len(self.df))
        self.tableWidget.setColumnCount(len(self.df.columns))
        self.tableWidget.setHorizontalHeaderLabels(self.df.columns)

        for i in range(len(self.df)):
            for j, column in enumerate(self.df.columns):
                value = str(self.df.iloc[i, j]) if pd.notna(self.df.iloc[i, j]) else ''
                item = QTableWidgetItem(value)
                self.tableWidget.setItem(i, j, item)

    def checkData(self):
        if self.df is None:
            self.logMessage("Ошибка: Данные не загружены!")
            return

        self.logText.show()
        self.logText.clear()

        try:
            # Этап 1: Нормализация данных
            try:
                self.df = self.processor.cleanData(self.df)  # Нормализация строк и дат
                self.logMessage("Данные нормализованы.")
                self.displayTable()  # Обновляем таблицу
            except Exception as e:
                self.logMessage(f"Ошибка нормализации данных: {e}")
                return


            # Этап 2: Проверка на наличие обязательных данных
            required_columns = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Место рождения', 'Регистрация',
                                'Организация', 'Должность']
            missing_data_info = {}

            for i in self.df.index:
                missing_columns = [col for col in required_columns if pd.isna(self.df.at[i, col])]
                if missing_columns:
                    missing_data_info[i] = missing_columns
                    self.df.at[i, 'Статус'] = f"Отсутствуют: {', '.join(missing_columns)}"  # Записываем статус

            # Логируем строки с отсутствующими данными
            for i, missing_columns in missing_data_info.items():
                self.logMessage(f"Строка {i + 1}: отсутствуют обязательные поля: {', '.join(missing_columns)}")

            # Обновляем таблицу после проверки обязательных данных
            self.displayTable()

            # Этап 3: Проверка на дублирование в файле
            try:
                self.df = self.processor.checkDuplicates(self.df)  # Обновляем DataFrame
                self.displayTable()  # Синхронизируем таблицу
            except Exception as e:
                self.logMessage(f"Ошибка проверки на дублирование: {e}")
                return

            # Этап 4: Проверка на дублирование в БД
            db_duplicate_indices = []
            for i in self.df.index:
                row = self.df.loc[i]
                surname = row['Фамилия']
                name = row['Имя']
                middle_name = row['Отчество']
                birth_date = row['Дата рождения']
                organization = row['Организация']

                if self.db_manager.find_matches(surname, name, middle_name, birth_date, organization):
                    db_duplicate_indices.append(i)
                    self.df.at[i, 'Статус'] = "Дубликат в БД"  # Добавляем статус

            # Обновляем таблицу после проверки дубликатов в БД
            self.displayTable()

            # Этап 5: Проверка подозрительных последовательностей дат
            suspicious_indices = []
            if 'Дата рождения' in self.df.columns:
                try:
                    suspicious_indices = self.processor.detectSequentialDates(self.df, 'Дата рождения')
                    for idx in suspicious_indices:
                        self.df.at[idx, 'Статус'] = "Подозрительная дата"  # Обновляем статус
                except Exception as e:
                    self.logMessage(f"Ошибка проверки последовательности дат: {e}")

            self.displayTable()  # Обновляем таблицу после проверки последовательностей

            # Этап 6: Цветовая маркировка
            for i in self.df.index:
                color = None
                status = self.df.at[i, 'Статус']
                if i in missing_data_info:
                    color = QColor(255, 165, 0)  # Оранжевый: отсутствуют данные
                elif self.df.at[i, 'Повтор']:
                    color = QColor(255, 0, 0)  # Красный: дублирование в файле
                elif i in db_duplicate_indices:
                    color = QColor(128, 0, 128)  # Фиолетовый: дублирование в БД
                elif i in suspicious_indices:
                    color = QColor(0, 0, 255)  # Синий: подозрительная последовательность дат
                else:
                    color = QColor(0, 255, 0)  # Зеленый: данные корректны
                    self.df.at[i, 'Добавлено'] = True

                for j in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(i, j) or QTableWidgetItem(str(self.df.iloc[i, j]))
                    item.setBackground(color)
                    self.tableWidget.setItem(i, j, item)


            self.uploadButton.setEnabled(True)  # Активируем кнопку добавления в БД
        except Exception as e:
            self.logMessage(f"Ошибка проверки данных: {e}")

    def uploadToDB(self):
        if self.df is None:
            self.logMessage("Ошибка: Данные не загружены!")
            return

        try:
            added_count = 0
            for index, row in self.df.iterrows():
                if row.get('Добавлено', True):  # Только строки со статусом "Добавлено"
                    self.db_manager.save_valid_data([row.to_dict()])
                    added_count += 1

            self.logMessage(f"Добавлено {added_count} строк в базу данных.")
            self.uploadButton.setEnabled(False)  # Отключаем кнопку после загрузки
        except Exception as e:
            self.logMessage(f"Ошибка загрузки данных в БД: {e}")

    def updateTableColors(self, invalid_indices, suspicious_indices):
        for i in range(self.tableWidget.rowCount()):
            color = None

            if i in invalid_indices:
                color = QColor(255, 165, 0)  # Оранжевый
            elif self.df.iloc[i].get('Повтор', False):
                color = QColor(255, 0, 0)  # Красный
            elif i in suspicious_indices:
                color = QColor(0, 0, 255)  # Синий
            elif self.df.iloc[i].get('Добавлено', False):
                color = QColor(0, 255, 0)  # Зеленый

            for j in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(i, j) or QTableWidgetItem(str(self.df.iloc[i, j]))
                if color:
                    item.setBackground(color)
                self.tableWidget.setItem(i, j, item)

    def saveFile(self):
        if self.df is None:
            self.logMessage("Ошибка: Нет данных для сохранения.")
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)")
        if not file_name:
            return

        try:
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Результаты проверки"

            color_map = {
                'Оранжевый': 'FFFFA500',
                'Красный': 'FFFF0000',
                'Зеленый': 'FF00FF00'
            }
            required_columns = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Место рождения', 'Регистрация',
                                'Организация', 'Должность']
            valid_data = self.df.dropna(subset=required_columns)

            # Запись данных в Excel
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = sheet.cell(row=r_idx + 1, column=c_idx + 1, value=value)

                    # Применяем цвета
                    if r_idx > 0:  # Пропускаем заголовки
                        table_row = r_idx - 1
                        if table_row in self.df.index:
                            if table_row in self.df[~self.df.index.isin(valid_data.index)].index:
                                cell.fill = PatternFill(start_color=color_map['Оранжевый'], fill_type="solid")
                            elif self.df.at[table_row, 'Повтор']:
                                cell.fill = PatternFill(start_color=color_map['Красный'], fill_type="solid")
                            elif self.df.at[table_row, 'Добавлено']:
                                cell.fill = PatternFill(start_color=color_map['Зеленый'], fill_type="solid")

            wb.save(file_name)
            self.logMessage(f"Файл сохранен: {file_name}")
        except Exception as e:
            self.logMessage(f"Ошибка сохранения файла: {e}")

    def addToBlacklist(self):
        """
        Добавляет выбранного пользователя в черный список.
        """
        try:
            selected_row = self.tableWidget.currentRow()
            # print(selected_row)
            if selected_row != -1:
                surname = self.tableWidget.item(selected_row, 1).text()
                name = self.tableWidget.item(selected_row, 2).text()
                middle_name = self.tableWidget.item(selected_row, 3).text() or ""
                birth_date = self.tableWidget.item(selected_row, 4).text()
                # birth_place = self.tableWidget.item(selected_row, 4).text()
                # registration = self.tableWidget.item(selected_row, 5).text()
                # organization = self.tableWidget.item(selected_row, 6).text()
                # position = self.tableWidget.item(selected_row, 7).text()

                self.db_manager.add_person_to_blacklist(
                    surname, name, middle_name, birth_date)
                self.logMessage(f"{surname} {name} {middle_name} добавлен в черный список.")
            else:
                self.logMessage("Ошибка: Не выбрана строка для добавления в черный список.")
        except Exception as e:
            self.logMessage(f"Ошибка добавления в черный список: {e}")


